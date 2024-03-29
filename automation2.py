import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import NamedStyle

wb = xl.load_workbook('accounts.xlsx')
sheet = wb['Creditors']
sheet1 = wb.create_sheet("Total")

min_row = sheet.min_row
max_row = sheet.max_row
min_col = sheet.min_column
max_col = sheet.max_column

for row in range(min_row, max_row+1):
    for col in range(min_col, max_col+1):
        sheet1.cell(row,col).value = sheet.cell(row,col).value
        # sheet1.cell(row,col).font = sheet.cell(row,col).font
        # sheet1.cell(row, col).value.font = sheet.cell(row, col).font
        # sheet1.cell(row, col).value.fill = sheet.cell(row, col).fill
        # sheet1.cell(row, col).value.border = sheet.cell(row, col).border
        # sheet1.cell(row, col).value.alignment = sheet.cell(row, col).alignment


for row in range(min_row+2, max_row+1):
    for col in range(min_col+6, max_col+1):
        sheet1.cell(row, col).value = 1000
        sheet1.cell(row,col).style = 'Currency'

 for row in range(min_row+2, max_row+1):
    for col in range(min_col+6, max_col+1):
        sheet.cell(max_row)

# for row in range(min_row+2, max_row+1):
#     sheet1.cell(row,max_col+1).value = 0
#
# for row in range(min_row+2, max_row+1):
#     for col in range(min_col+6, max_col+1):
#         sheet1.cell(row,max_col+1).value += sheet1.cell(row,col).value
#         sheet1.cell(row,max_col+1).style = 'Currency'

wb.save('accounts3.xlsx')





