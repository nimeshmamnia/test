import openpyxl as xl
from openpyxl.styles import Font

wb = xl.load_workbook('accounts.xlsx')
sheet = wb['Creditors']
sheet1 = wb.create_sheet("Total")

min_row = sheet.min_row
max_row = sheet.max_row
min_col = sheet.min_column
max_col = sheet.max_column

# Copy cell values and apply font formatting
for row in range(min_row, max_row + 1):
    for col in range(min_col, max_col + 1):
        source_cell = sheet.cell(row=row, column=col)
        destination_cell = sheet1.cell(row=row, column=col, value=source_cell.value)
        destination_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            # italic=source_cell.font.italic,
            # color=source_cell.font.color,
            # underline=source_cell.font.underline,
            # strike=source_cell.font.strike,
        )

# Apply currency style to specified columns
for row in range(min_row + 2, max_row + 1):
    for col in range(min_col + 6, max_col + 1):
        sheet1.cell(row=row, column=col, value=1000)
        sheet1.cell(row=row, column=col).number_format = 'Currency'

# Calculate and format total column
for row in range(min_row + 2, max_row + 1):
    total = 0
    for col in range(min_col + 6, max_col + 1):
        total += sheet1.cell(row=row, column=col).value
    sheet1.cell(row=row, column=max_col + 1, value=total)
    sheet1.cell(row=row, column=max_col + 1).number_format = 'Currency'

wb.save('accounts4.xlsx')
