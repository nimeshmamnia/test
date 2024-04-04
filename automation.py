import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('accounts.xlsx')
sheet = wb['Creditors']

for row in range(3, sheet.max_row + 1):
    for column in range(7, sheet.max_column + 1):
        sheet.cell(row, column).value = 1000

for row in range(3, sheet.max_row + 1):
    sheet.cell(row, 6).value = 0

for row in range(3, sheet.max_row + 1):
    for column in range(7, sheet.max_column + 1):
        sheet.cell(row, 6).value += sheet.cell(row, column).value

ref = Reference(sheet, min_row=3, max_row=sheet.max_row, min_col=6, max_col=6)

chart = BarChart()
chart.add_data(ref)
sheet.add_chart(chart, 'h30')
wb.save('accounts2.xlsx')
